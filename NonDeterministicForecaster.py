import os
import json
import argparse
from typing import Dict, Any, List, Optional
import openpyxl
from cerebras.cloud.sdk import Cerebras
from docx import Document  # pip install python-docx
import time
import logging
import random
import re

#This is a non-deterministic forecaster version of the main model. This version has the task of building the formulas itself. 
# This is more flexible but also quality can vary. Use at your own risk (or just have fun)!

# ---------------- User configurable paths ----------------
SOURCE_FILE = "YOUR_FILEest.xlsx"  # primary data workbook to read and write

# Hard-coded row/label mapping (aliases all point to the same row number)
LABEL_TO_ROW: Dict[str, int] = {
    # Revenue
    "revenue": 5,
    "revenues": 5,

    # COGS / cost of sales
    "cogs": 6,
    "cost of goods sold": 6,

    # Gross profit
    "gross profit": 7,

    # SG&A
    "sg&a": 8,
    "sg and a": 8,
    "selling, general & administrative": 8,

    # D&A
    "d&a": 9,
    "depreciation & amortization": 9,

    # Interest
    "interest income": 10,
    "interest expense": 11,

    # Profit before tax / taxes / net income
    "profit before tax": 12,
    "profit before taxes": 12,
    "tax expense": 13,
    "tax expenses": 13,
    "net income": 14,
}

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template structure extraction
# ---------------------------------------------------------------------------
def extract_template_structure(template_path: str) -> Dict[str, str]:
    """Extract row labels and their cell coordinates from template."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]
    structure = {}

    # Look for row labels in column A (assuming labels are in column A)
    for row in range(1, 50):  # Check first 50 rows
        cell = ws[f"A{row}"]
        if cell.value and isinstance(cell.value, str) and cell.value.strip():
            label = cell.value.strip()
            # Map to the data column (E) for the same row
            structure[label] = f"E{row}"

    logger.info("Extracted template structure: %s", structure)
    return structure

# ---------------------------------------------------------------------------
# Global Cerebras client & model configuration
# ---------------------------------------------------------------------------
CLIENT = Cerebras(api_key=os.environ.get("CEREBRAS_API_KEY"))
MODEL = "qwen-3-coder-480b"

# ---------------------------------------------------------------------------
# Spreadsheet helpers
# ---------------------------------------------------------------------------
def load_excel_data(filepath: str) -> Dict[str, List[List[Any]]]:
    """Load every worksheet in an Excel workbook into a matrix list-of-lists."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data: Dict[str, List[List[Any]]] = {}
    for ws in wb.worksheets:
        sheet_matrix = [list(row) for row in ws.iter_rows(values_only=True)]
        data[ws.title] = sheet_matrix
        logger.debug("Loaded sheet '%s' with %d rows", ws.title, len(sheet_matrix))
    return data

# ---------------------------------------------------------------------------
# Cerebras wrapper with retry/backoff
# ---------------------------------------------------------------------------
def _chat_with_retries(messages: List[Dict[str, Any]], *, purpose: str, max_retries: int = 3, base_delay: float = 0.5) -> str:
    """Call Cerebras chat API with exponential-backoff retries."""
    last_error: Exception | None = None
    for attempt in range(1, max_retries + 1):
        try:
            logger.info("Cerebras call (%s) attempt %d/%d", purpose, attempt, max_retries)
            response = CLIENT.chat.completions.create(model=MODEL, messages=messages)
            return response.choices[0].message.content
        except Exception as e:
            last_error = e
            if attempt == max_retries:
                break
            sleep_s = base_delay * (2 ** (attempt - 1)) + random.uniform(0, 0.25)
            logger.warning("Cerebras call failed (%s) on attempt %d: %s – retrying in %.2fs", purpose, attempt, e, sleep_s)
            time.sleep(sleep_s)
    logger.error("Cerebras call failed after %d attempts (%s)", max_retries, purpose)
    raise last_error if last_error else RuntimeError("Unknown Cerebras error")

# ---------------------------------------------------------------------------
# AI-assisted analytics steps
# ---------------------------------------------------------------------------
def identify_assumptions(workbook_data: Dict[str, Any], core_elements: Optional[str] = None, question: Optional[str] = None) -> str:
    """Cerebras call – extract numeric data & assumptions from raw workbook."""
    system_prompt = (
        "You are a financial analyst. Given the core elements and raw workbook data for P&L forecasting. "
        "extract specific financial information and all underlying assumptions required for modelling."
        "Do not do any modeling, and prepare all potentially relevant information for modeling (P&L)"
    )
    context = {
        "workbook": workbook_data,
        "core_elements": core_elements,
        "question": question or "Analyze the financial performance and provide key insights",
    }
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(context)},
    ]
    return _chat_with_retries(messages, purpose="identify_assumptions")


def make_formulas(workbook_data: Dict[str, Any], template_structure: Dict[str, str], core_elements: Optional[str] = None, question: Optional[str] = None) -> str:
    """Cerebras call – return row-based calculations using template structure."""
    extra_rules = f"\nCalculation rules (JSON):\n{core_elements}\n" if core_elements else ""

    system_prompt = (
        "You are a senior financial analyst creating a 5-year financial model. "
        f"Use this template structure: {template_structure}\n\n"
        f"{extra_rules}"
        "Return a JSON array where each object has: row_label, values.\n"
        "The **values** field must be an array with FIVE entries – one for each forecast year 2023-2027, in order.\n\n"
        "• Each entry may be a literal number or, for derived rows, an arithmetic expression built only from numeric literals and + - * / ( ).\n"
        "• Do NOT reference other cells and do NOT evaluate the expressions yourself.\n\n"
        "Example object:\n"
        "  {\"row_label\": \"Revenue\", \"values\": [1000, \"1000*1.05\", \"1000*1.05*1.05\", \"...\", \"...\"]}\n\n"
        "IMPORTANT: Include all rows: Revenue, COGS, Gross Profit, Gross Margin %, SG&A, D&A, EBIT, EBIT Margin %, Interest Income, Interest Expense, Profit Before Tax, Tax Expense, Net Income.\n"
        "Return ONLY valid JSON. Do not add commentary."
    )
    context = {
        "workbook": workbook_data,
        "template_structure": template_structure,
        "question": question or "Create a 5-year financial forecast",
    }
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(context)},
    ]
    return _chat_with_retries(messages, purpose="generate_formulas")


# ---------------------------------------------------------------------------
# Formula evaluation engine
# ---------------------------------------------------------------------------
def values_from_json(json_text: str, row_mapping: Dict[str, int] = LABEL_TO_ROW) -> Dict[str, Dict[str, float]]:
    """Convert row-based JSON into cell-based nested dict using the label alias mapping."""
    import json, numbers, re

    data = json.loads(json_text)
    out: Dict[str, Dict[str, float]] = {}

    forecast_columns = ["E", "F", "G", "H", "I"]  # 5-year span

    # create a normalised lookup once so aliases with punctuation map correctly
    normalised_map = {re.sub(r"[^a-z% ]", "", k.lower()).strip(): v for k, v in row_mapping.items()}

    for obj in data:
        raw_label = obj.get("row_label")
        vals = obj.get("values")

        if raw_label is None or vals is None:
            print(f"Skipping object missing fields: {obj}")
            continue

        if len(vals) != 5:
            print(f"Expected 5 values for {raw_label}, got {len(vals)} – skipping")
            continue

        # normalise label for lookup
        label = re.sub(r"[^a-z% ]", "", raw_label.lower()).strip()
        if label not in normalised_map:
            print(f"Skipped unknown row label: {raw_label}")
            continue

        row_num = normalised_map[label]
        sheet = "Income Statement"

        for i, col in enumerate(forecast_columns):
            val = vals[i]
            try:
                if isinstance(val, str):
                    expr = val.replace("^", "**")  # allow power operator
                    print(f"Evaluating expression for {raw_label} {col} (row {row_num}): {expr}")
                    num = float(eval(expr, {"__builtins__": {}}))
                elif isinstance(val, numbers.Number):
                    num = float(val)
                else:
                    print(f"Unsupported value type for {raw_label} {col}: {val}")
                    continue
            except Exception as e:
                print(f"Failed to eval {raw_label} {col}: {e}; skipping cell")
                continue

            cell = f"{col}{row_num}"
            out.setdefault(sheet, {})[cell] = num

    print(f"Completed values_from_json; total cells processed: {sum(len(d) for d in out.values())}")
    return out


def fill_template(values: Dict[str, Dict[str, float]], template_path: str, output_path: Optional[str] = None) -> str:
    """Insert evaluated numbers into an Excel template.

    Args:
        values: Nested dict from evaluate_formulas – sheet -> cell -> number.
        template_path: Path to the template workbook (e.g. incomestatementformat.xlsx).
        output_path: Where to save the filled workbook. If None, writes alongside template with suffix _filled.xlsx.

    Returns:
        The path to the written workbook.
    """
    from pathlib import Path

    wb = openpyxl.load_workbook(template_path)

    missing_refs = []
    for sheet_name, cell_map in values.items():
        if sheet_name not in wb.sheetnames:
            missing_refs.append((sheet_name, None))
            continue  # skip unknown sheet
        ws = wb[sheet_name]
        for cell, num in cell_map.items():
            try:
                ws[cell]  # trigger coordinate parsing
            except ValueError:
                missing_refs.append((sheet_name, cell))
                continue
            ws[cell] = num

    if missing_refs:
        logger.warning("Skipped %d references not present in template: %s", len(missing_refs), missing_refs[:20])

    if output_path is None:
        p = Path(template_path)
        output_path = str(p.with_name(p.stem + "_filled" + p.suffix))

    wb.save(output_path)
    logger.info("Template filled and saved to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Template sheet handling
# ---------------------------------------------------------------------------
def copy_template_sheet(template_path: str, dest_path: str, new_sheet_name: Optional[str] = None) -> str:
    """Copy the first sheet from template workbook into destination workbook.

    Overwrites any existing sheet with the same name.
    Returns the sheet name used.
    """
    tpl_wb = openpyxl.load_workbook(template_path, data_only=True)
    tpl_ws = tpl_wb.worksheets[0]
    target_name = new_sheet_name or tpl_ws.title

    dest_wb = openpyxl.load_workbook(dest_path)
    # Remove existing sheet with that name to ensure fresh copy
    if target_name in dest_wb.sheetnames:
        std = dest_wb[target_name]
        dest_wb.remove(std)

    new_ws = dest_wb.create_sheet(title=target_name)

    # Copy dimensions (column widths)
    for col_dim in tpl_ws.column_dimensions.values():
        new_ws.column_dimensions[col_dim.column_letter].width = col_dim.width
    for row_dim in tpl_ws.row_dimensions.values():
        new_ws.row_dimensions[row_dim.index].height = row_dim.height

    # Copy cell values and basic styles
    from copy import copy as _copy
    for row in tpl_ws.iter_rows():
        for cell in row:
            new_cell = new_ws[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = _copy(cell.font)
                new_cell.border = _copy(cell.border)
                new_cell.fill = _copy(cell.fill)
                new_cell.number_format = _copy(cell.number_format)
                new_cell.protection = _copy(cell.protection)
                new_cell.alignment = _copy(cell.alignment)

    dest_wb.save(dest_path)
    logger.info("Copied template sheet '%s' into %s", target_name, dest_path)
    return target_name


def populate_template(values: Dict[str, float], dest_path: str, sheet_name: str) -> None:
    """Write numbers into existing sheet, only where coordinates already exist."""
    wb = openpyxl.load_workbook(dest_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} not found in {dest_path}")
    ws = wb[sheet_name]

    for cell, num in values.items():
        try:
            ws[cell].value = num
        except ValueError:
            logger.debug("Skipping invalid cell coordinate %s", cell)

    wb.save(dest_path)
    logger.info("Populated %d values into sheet '%s'", len(values), sheet_name)


# ---------------------------------------------------------------------------
# Formula evaluation engine
# ---------------------------------------------------------------------------


# new function for inserting into any workbook
def insert_values_sheet(values: Dict[str, float], workbook_path: str, sheet_title: str) -> None:
    """Create a new sheet and write cell->value mappings.

    Args:
        values: Flat dict of cell address -> number (e.g. {"A1": 10}).
        workbook_path: Existing workbook file to modify.
        sheet_title: Title for the new sheet.
    """
    wb = openpyxl.load_workbook(workbook_path)
    if sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
        logger.info("Overwriting existing sheet '%s' in %s", sheet_title, workbook_path)
    else:
        ws = wb.create_sheet(title=sheet_title)

    for cell, num in values.items():
        ws[cell] = num

    wb.save(workbook_path)
    logger.info("Inserted %d values into sheet '%s' of %s", len(values), sheet_title, workbook_path)


# ---------------------------------------------------------------------------
# Final AI review helper
# ---------------------------------------------------------------------------
def last_check(workbook_values: Dict[str, Dict[str, float]], summary_question: Optional[str] = None) -> str:
    """Ask the model to review the filled data and flag any issues.

    Args:
        workbook_values: Nested dict of the numbers we plan to write – typically the result of evaluate_formulas.
        summary_question: Optional extra question/prompts for the model.
    Returns:
        Model response containing either 'All good' or suggestions/corrections.
    """
    system_prompt = (
        "You are a formatting auditor. Examine the structured workbook data intended for presentation. "
        "Check for alignment with standard financial-model formatting: correct headers, year columns, sourceline totals, consistent decimal places, and logical ordering. "
        "If everything is presentation-ready, respond with 'PASS'. Otherwise, list the formatting issues and specify the exact cell edits required to fix them (sheet, cell, new_value or action)."
    )
    context = {
        "values": workbook_values,
        "question": summary_question or "Perform a final sanity check on the financial model output."
    }
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(context)},
    ]
    return _chat_with_retries(messages, purpose="final_review")

# ---------------------------------------------------------------------------
# Formatting auto-fix helper
# ---------------------------------------------------------------------------

def apply_formatting_fixes(review_text: str, workbook_path: str, sheet_name: str = "Income Statement") -> None:
    """Parse the LLM review for simple cell-level edits and apply them.

    Expected edit lines of the form:
      • SheetName | Cell | NewValue
    or
      • Cell | NewValue   (sheet inferred)

    Anything that cannot be parsed is ignored silently.
    If you see any data that is clearly wrong on the projection page, such as a missed decimal place, correct that too.
    This is deliberately conservative so the model can only tweak explicit cells.
    """
    import re
    edits = []
    pattern = re.compile(r"(?P<sheet>[A-Za-z0-9 _]+)?\s*\|\s*(?P<cell>[A-Za-z]+[0-9]+)\s*\|\s*(?P<value>.+)")
    for line in review_text.splitlines():
        m = pattern.search(line)
        if m:
            sheet = m.group("sheet") or sheet_name
            cell = m.group("cell").strip()
            new_val = m.group("value").strip()
            edits.append((sheet, cell, new_val))

    if not edits:
        print("No actionable formatting edits parsed from review.")
        return

    wb = openpyxl.load_workbook(workbook_path)
    applied = 0
    for sheet, cell, new_val in edits:
        if sheet not in wb.sheetnames:
            print(f"Skip unknown sheet {sheet} in edit {sheet}|{cell}")
            continue
        ws = wb[sheet]
        try:
            # attempt to convert numeric strings to float
            if re.fullmatch(r"-?\d+(\.\d+)?", new_val):
                ws[cell].value = float(new_val)
            else:
                ws[cell].value = new_val
            applied += 1
        except ValueError:
            print(f"Invalid cell coordinate in edit {sheet}|{cell}")

    if applied:
        wb.save(workbook_path)
        print(f"Applied {applied} formatting fixes as suggested by the model.")


# ---------------------------------------------------------------------------
# Command-line entrypoint
# ---------------------------------------------------------------------------
def main() -> None:
    """Run the forecasting pipeline writing results back into the source workbook."""
    user_question = input("Enter analysis question (or press Enter for default): ") or None

    wb_data = load_excel_data(SOURCE_FILE)

    # Load modeling rules so the LLM follows strict sign discipline
    core_rules: Optional[str] = None
    try:
        with open("instructions.json", "r") as f:
            core_rules = f.read()
    except Exception:
        core_rules = None

    print("\n--- Calling identify_assumptions ---")
    assumptions = identify_assumptions(wb_data, core_rules, user_question)
    print("AI assumptions response:\n", assumptions)

    print("\n--- Calling make_formulas ---")
    template_structure = extract_template_structure("incomestatementformat.xlsx")
    formulas_json = make_formulas(wb_data, template_structure, core_rules, user_question)
    print("Raw formulas JSON:\n", formulas_json)

    values = values_from_json(formulas_json)
    print("\nParsed numeric values:")
    print(values)

    # Ensure template sheet exists before any potential formatting edits
    tmpl_sheet = copy_template_sheet("incomestatementformat.xlsx", SOURCE_FILE, "Income Statement")

    review = last_check(values)
    print("\nAI review:\n", review)

    if "PASS" not in review.upper():
        apply_formatting_fixes(review, SOURCE_FILE, tmpl_sheet)
        print("Re-ran formatting fixes based on model suggestions.")

    # Merge all values and populate
    combined: Dict[str, float] = {}
    for sheet_dict in values.values():
        combined.update(sheet_dict)

    populate_template(combined, SOURCE_FILE, tmpl_sheet)
    print("\nUpdated", SOURCE_FILE, "with formatted sheet", tmpl_sheet)


if __name__ == "__main__":
    main()