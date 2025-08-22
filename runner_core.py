"""runner_core.py
Core utilities for the AI-driven finance runner.
Contains only generic, dependency-free helpers so other scripts can import
without pulling in heavy code.
"""
from __future__ import annotations

from typing import Dict, List, Any

import openpyxl
# import json
# import os
import re
import math
import functools
import operator


def load_all_sheets(path: str) -> Dict[str, List[List[Any]]]:
    """Read *every* sheet in an Excel workbook and return JSON-ready matrices.

    The outer dict maps sheet names -> 2-D list (rows of cell values). All cell
    objects are converted to their raw values via *values_only* mode.  Empty
    sheets are included with an empty list so the caller can detect them.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_data: Dict[str, List[List[Any]]] = {}
    for ws in wb.worksheets:
        # Convert generator to list so we can reuse multiple times
        matrix: List[List[Any]] = [list(row) for row in ws.iter_rows(values_only=True)]
        sheet_data[ws.title] = matrix
    return sheet_data


# Convenience re-export so callers can keep old name
load_selected_sheets = load_all_sheets


# ---------------------------------------------------------------------------
# Assumption extractor (flattened list for logging / prompt context)
# ---------------------------------------------------------------------------


def _find_year_headers(ws) -> dict[int, int]:
    """Scan header area for year headers; accept ints or numeric strings.
    Returns {year: col_idx} when >=3 years are detected.
    """
    def parse_year(v) -> int | None:
        if isinstance(v, (int, float)):
            y = int(v)
            return y if 2000 <= y <= 2100 else None
        if isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                y = int(s)
                return y if 2000 <= y <= 2100 else None
        return None

    max_scan_rows = min(40, ws.max_row)
    # Primary: row with 3+ detectable year headers
    for header_row in range(1, max_scan_rows + 1):
        mapping: dict[int, int] = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=col).value
            y = parse_year(v)
            if y is not None:
                mapping[y] = col
        if len(mapping) >= 3:
            return mapping

    # Fallback: find a single 2022 header and assume contiguous years to the right
    for header_row in range(1, max_scan_rows + 1):
        for col in range(1, ws.max_column + 1):
            y = parse_year(ws.cell(row=header_row, column=col).value)
            if y == 2022:
                mapping: dict[int, int] = {}
                year = 2022
                c = col
                # Map forward while within sheet bounds
                while c <= ws.max_column and year <= 2100:
                    mapping[year] = c
                    year += 1
                    c += 1
                if len(mapping) >= 3:
                    return mapping
    return {}


def list_assumptions(path: str, sheet_name: str = "Assumptions") -> list[dict[str, Any]]:
    """Return flattened assumption rows as list[{name, year, value}].

    Attempts multiple likely label columns (C, B, A, D) and accepts year headers
    as ints or numeric strings within the first 40 rows.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    year_cols = _find_year_headers(ws)
    if not year_cols:
        return []

    # Try several label columns; use the first that yields data
    for label_col in (3, 2, 1, 4):  # C, B, A, D
        collected: list[dict[str, Any]] = []
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=label_col).value
            if not isinstance(label, str) or not label.strip():
                continue
            label = label.strip()
            for y, col in year_cols.items():
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                try:
                    collected.append({"name": label, "year": int(y), "value": float(val)})
                except Exception:
                    continue
        if collected:
            return sorted(collected, key=lambda d: (d["name"], d["year"]))
    return []


def assumptions_debug(path: str, sheet_name: str = "Assumptions") -> dict[str, Any]:
    """Return diagnostics for assumptions extraction to aid debugging."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    result: Dict[str, Any] = {"sheets": sheets, "sheet_exists": sheet_name in sheets}
    if sheet_name not in sheets:
        return result
    ws = wb[sheet_name]
    # sample first 8 rows
    sample = []
    for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True):
        sample.append(list(row))
    result["sample_rows"] = sample
    # header detection
    result["year_cols"] = _find_year_headers(ws)
    # column tries
    tries: Dict[str, int] = {}
    for label_col in (3, 2, 1, 4):
        count = 0
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=label_col).value
            if not isinstance(label, str) or not label.strip():
                continue
            for _y, col in result["year_cols"].items():
                val = ws.cell(row=r, column=col).value
                if val is not None:
                    count += 1
        tries[f"col_{label_col}"] = count
    result["label_column_trials"] = tries
    return result


# ---------------------------------------------------------------------------
# Cerebras helper (step 3)
# ---------------------------------------------------------------------------


def call_cerebras_for_plan(
    workbook_json: dict[str, Any],
    question: str,
    *,
    api_key: str | None = None,
    mode: str = "dsl",  # "dsl" or "python"
) -> dict[str, Any]:
    """Send workbook + question to Cerebras and return the JSON plan.

    mode="dsl"  -> expect {"answer_label", "formulas": [...]}.
    mode="python" -> expect {"answer_label", "python_code": "def compute(...): ..."}.
    """
    try:
        from cerebras.cloud.sdk import Cerebras  # type: ignore
    except ImportError as e:  # pragma: no cover
        raise ImportError("cerebras-cloud-sdk not installed") from e

    client = Cerebras(api_key=api_key or os.environ.get("CEREBRAS_API_KEY"))

    if mode not in ("dsl", "python"):
        raise ValueError("mode must be 'dsl' or 'python'")

    if mode == "dsl":
        system_prompt = (
            "You are a financial-model assistant. Use the workbook JSON to build "
            "a formula plan. Return STRICT JSON with keys 'answer_label', 'series_values', and 'formulas'. "
            "'series_values' must be a map of assumption label -> {year:number}. For example: "
            "{ 'Revenue (Yoy.)': {'2023':0.05,'2024':0.05}, 'Tax': {'2023':0.28} }. "
            "The 'formulas' value MUST be an array of JSON OBJECTS (not strings, not code). "
            "Each object MUST have exactly two keys: 'row_label' (string) and 'expr' (object). "
            "The 'expr' is JSON using ONLY these terminals and ops: "
            "terminals = { 'const': number } or { 'ref': string } or { 'series': string } or { 'op':'lag', 'ref': string, 'periods': number } ; "
            "ops = { 'op':'add'|'sub'|'mul'|'div', 'args':[<expr>, ...] } or { 'op':'neg', 'arg': <expr> }. "
            "'ref' takes ONLY a label string (no commas, no ,0). Use lag for prior-year. "
            "ABSOLUTELY FORBIDDEN: strings like \"series('X', mul(...))\" or function-like syntax; only pure JSON objects. "
            "Examples (MUST MATCH EXACT SHAPE): "
            "- (5 - 3): { 'row_label':'Example Sub', 'expr': { 'op':'sub', 'args':[ {'const':5}, {'const':3} ] } } "
            "- (2 + 9): { 'row_label':'Example Add', 'expr': { 'op':'add', 'args':[ {'const':2}, {'const':9} ] } } "
            "- Revenue * (1 + Revenue YoY): { 'row_label':'Revenue', 'expr': { 'op':'mul', 'args':[ { 'ref':'Revenue' }, { 'op':'add', 'args':[ { 'const':1 }, { 'series':'Revenue (Yoy.)' } ] } ] } } "
            "- Prior-year EBIT: { 'row_label':'EBIT', 'expr': { 'op':'lag', 'ref':'EBIT', 'periods':1 } } "
            "Return ONLY JSON (no markdown)."
        )
    else:
        system_prompt = (
            "You are a financial-model assistant. Use the workbook JSON to write Python "
            "under key 'python_code'. The code must define compute(sheet_data, assumptions) "
            "and return a dict {row_label: {year: value}} for 2023-2027. No imports except math. "
            "Return STRICT JSON with keys 'answer_label' and 'python_code'."
        )

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps({"workbook": workbook_json, "question": question})},
    ]

    resp = client.chat.completions.create(model="qwen-3-coder-480b", messages=messages)
    content = resp.choices[0].message.content  # type: ignore[attr-defined]
    try:
        return json.loads(content)
    except json.JSONDecodeError as e:
        raise ValueError(f"Model did not return valid JSON: {e}\nReceived:\n{content[:500]}")


# ---------------------------------------------------------------------------
# Minimal DSL evaluator (step 4)
# ---------------------------------------------------------------------------


def _build_row_map(operating_matrix: list[list[Any]], label_col: int = 2) -> dict[str, int]:
    """Return {label: row_index} for Operating Model sheet (0-based row index)."""
    mapping: dict[str, int] = {}
    for idx, row in enumerate(operating_matrix):
        if label_col < len(row) and isinstance(row[label_col], str):
            mapping[row[label_col].strip()] = idx
    return mapping

# basic synonyms (extend as needed)
_SYNONYMS = {
    "net inc.": "Net Income",
    "net income": "Net Income",
    "gross profit": "Gross Profit",
    "ebit": "EBIT",
}

def _normalize_label(lbl: str) -> str:
    key = lbl.strip().lower()
    return _SYNONYMS.get(key, lbl)


def _detect_year_cols(matrix: list[list[Any]], max_header_rows: int = 20) -> dict[int, int]:
    """Scan first rows of a matrix for numeric year headers; return {year: col_index}."""
    for r in range(min(max_header_rows, len(matrix))):
        mapping: dict[int, int] = {}
        for c, val in enumerate(matrix[r]):
            if isinstance(val, (int, float)) and 2000 <= int(val) <= 2100:
                mapping[int(val)] = c
        if len(mapping) >= 3:
            return mapping
    return {}


def evaluate_dsl(
    formulas_plan: dict[str, Any],
    workbook_json: dict[str, Any],
    assumptions: list[dict[str, Any]],
    years: tuple[int, ...] = (2023, 2024, 2025, 2026, 2027),
) -> dict[str, Any]:
    """Evaluate DSL formulas into numeric updates.

    Returns {"updates": [ {row_label, year_values} ] }
    """
    if "formulas" not in formulas_plan:
        return {"updates": []}

    # Fallback: handle string formulas like "Label: ... = 123.45'" by parsing the RHS number
    formulas_raw = formulas_plan.get("formulas")
    if isinstance(formulas_raw, list) and formulas_raw and all(isinstance(f, str) for f in formulas_raw):
        parsed: dict[str, dict[int, float]] = {}
        # Capture label before ':' and numeric RHS after '='; allow commas and optional trailing quote
        rhs_re = re.compile(r"^\s*([^:]+):.*?=\s*([+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?)\s*(?:['\"])??\s*$")
        for line in formulas_raw:
            m = rhs_re.match(line)
            if not m:
                continue
            label = m.group(1).strip()
            num_str = m.group(2).replace(",", "")
            try:
                val = float(num_str)
            except Exception:
                continue
            parsed.setdefault(label, {})[years[0]] = val
        updates_list = []
        for lbl, by_year in parsed.items():
            year_values = {str(y): by_year.get(y) for y in years if y in by_year}
            if year_values:
                updates_list.append({"row_label": lbl, "year_values": year_values})
        return {"updates": updates_list}

    op_sheet = workbook_json.get("Operating Model")
    if op_sheet is None:
        return {"updates": []}

    row_map = _build_row_map(op_sheet)
    year_cols = _detect_year_cols(op_sheet)
    if 2022 not in year_cols:
        return {"updates": []}

    # Build assumptions dict of series-> {year: value}
    series_map: dict[str, dict[int, float]] = {}
    for item in assumptions:
        series_map.setdefault(item["name"], {})[item["year"]] = item["value"]
    # Merge any AI-provided series_values directly (full dump from sheet)
    if isinstance(formulas_plan, dict) and isinstance(formulas_plan.get("series_values"), dict):
        for key, per_year in formulas_plan["series_values"].items():
            for y_str, val in per_year.items():
                try:
                    y = int(y_str)
                    series_map.setdefault(key, {})[y] = float(val)
                except Exception:
                    continue

    # Seed 2022 base values from operating sheet
    values: dict[str, dict[int, float]] = {}
    col2022 = year_cols[2022]
    for label, r in row_map.items():
        if r < len(op_sheet) and col2022 < len(op_sheet[r]):
            v = op_sheet[r][col2022]
            if isinstance(v, (int, float)):
                values.setdefault(label, {})[2022] = float(v)

    def get_series(name: str, year: int):
        return series_map.get(name, {}).get(year)

    def eval_expr(expr: Any, year: int):
        if isinstance(expr, dict):
            if "const" in expr:
                return float(expr["const"])
            if "ref" in expr:
                val = values.get(str(expr["ref"]), {}).get(year)
                return val
            if "series" in expr:
                return get_series(str(expr["series"]), year)
            op = expr.get("op")
            if op == "lag":
                ref = str(expr.get("ref"))
                periods = int(expr.get("periods", 1))
                return values.get(ref, {}).get(year - periods)
            # collect arg values
            if op == "neg":
                inner = eval_expr(expr.get("arg"), year)
                return -inner if inner is not None else None

            args = [eval_expr(a, year) for a in expr.get("args", [])]
            if any(a is None for a in args):
                return None

            DISPATCH = {
                "add": lambda xs: sum(xs),
                "sub": lambda xs: xs[0] - sum(xs[1:]),
                "mul": lambda xs: math.prod(xs),
                "div": lambda xs: functools.reduce(operator.truediv, xs),
            }
            func = DISPATCH.get(op)
            return func(args) if func else None
        return None

    formulas = formulas_plan["formulas"]
    for _ in range(5):
        changed = False
        for f in formulas:
            if not isinstance(f, dict):
                continue
            label_raw = f.get("row_label")
            label = _normalize_label(label_raw) if isinstance(label_raw, str) else ""
            expr = f.get("expr")
            if not label or expr is None:
                continue
            for y in years:
                new_v = eval_expr(expr, y)
                cur_v = values.get(label, {}).get(y)
                if new_v is not None and (cur_v is None or abs(new_v - cur_v) > 1e-9):
                    values.setdefault(label, {})[y] = new_v
                    changed = True
        if not changed:
            break

    updates = []
    for label, by_year in values.items():
        yr_vals = {str(y): by_year.get(y) for y in years if y in by_year}
        if yr_vals:
            updates.append({"row_label": label, "year_values": yr_vals})
    return {"updates": updates}


# ---------------------------------------------------------------------------
# Simple validator (Step 5)
# ---------------------------------------------------------------------------


def validate_balance_sheet( #Maybe not needed, but we'll keep it in here
    updates: dict[str, Any],
    *,
    required_labels: tuple[str, str] = ("Total Assets", "Total Equity and Liabilities"),
    tolerance: float = 1e-2,
) -> tuple[bool, dict[str, Any]]:
    """Check that Total Assets equals Total Equity+Liabilities per year.

    Returns (is_ok, report_dict).
    If a year is missing from either side, it is ignored.
    """
    ta_label, tel_label = required_labels
    assets = next((u for u in updates.get("updates", []) if u.get("row_label") == ta_label), None)
    equity = next((u for u in updates.get("updates", []) if u.get("row_label") == tel_label), None)
    report: dict[str, Any] = {}
    ok = True
    if not assets or not equity:
        report["error"] = "Missing required labels for balance check"
        return False, report

    for y_str, a_val in assets.get("year_values", {}).items():
        e_val = equity.get("year_values", {}).get(y_str)
        if a_val is None or e_val is None:
            continue
        diff = abs(a_val - e_val)
        report[y_str] = diff
        if diff > tolerance:
            ok = False
    return ok, report