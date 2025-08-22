"""End-to-end pipeline for offline spreadsheet processing.

1. Loader – reads Excel or CSV into a matrix and an address map.
2. AiParser – calls Cerebras Cloud SDK to convert natural-language instructions into
   structured operations that reference addresses.
3. LogicEngine – executes operations on the matrix.
4. Exporter – writes the updated matrix back to CSV or Excel.

This file implements thin, testable components so the heavy lifting happens in Python
rather than inside Excel.
"""

from __future__ import annotations

import os
import csv
import json
from typing import List, Dict, Tuple, Any, Iterable, Union

import openpyxl

try:
    # Cerebras SDK is optional at import time – only required when AiParser is used.
    from cerebras.cloud.sdk import Cerebras  # type: ignore
except ImportError:  # pragma: no cover
    Cerebras = None  # type: ignore

Matrix = List[List[Any]]
AddressMap = Dict[Tuple[int, int], str]  # (row_idx, col_idx) -> "A1"


class SpreadsheetLoader:
    """Load a spreadsheet (CSV or Excel) into a matrix + address map."""

    def __init__(self, path: str):
        self.path = path
        self.matrix: Matrix = []
        self.address_map: AddressMap = {}

    def load(self) -> None:
        if self.path.lower().endswith(".csv"):
            self._load_csv()
        else:
            self._load_excel()

    # ---------------------------------------------------------------------
    # Internal helpers
    # ---------------------------------------------------------------------
    def _load_csv(self) -> None:
        """Simple CSV reader. Addresses are computed manually (A1 style)."""
        with open(self.path, newline="", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            for r_idx, row in enumerate(reader):
                self.matrix.append(row)
                for c_idx, _ in enumerate(row):
                    self.address_map[(r_idx, c_idx)] = self._idx_to_address(r_idx, c_idx)

    def _load_excel(self) -> None:
        """Use openpyxl so we get accurate cell coordinates."""
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = wb.active  # default to first sheet for now
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_vals: List[Any] = []
            for c_idx, value in enumerate(row):
                cell_address = ws.cell(row=r_idx + 1, column=c_idx + 1).coordinate
                row_vals.append(value)
                self.address_map[(r_idx, c_idx)] = cell_address
                print(value)
            self.matrix.append(row_vals)

    # Static ----------------------------------------------------------------
    @staticmethod
    def _idx_to_address(r: int, c: int) -> str:
        """Convert zero-based (row, col) to Excel address like A1."""
        col = ""
        n = c
        while True:
            n, remainder = divmod(n, 26)
            col = chr(65 + remainder) + col
            if n == 0:
                break
            n -= 1
        return f"{col}{r + 1}"


class AiParser:
    """Thin wrapper around Cerebras Cloud SDK call."""

    MODEL_NAME = "qwen-3-coder-480b"

    def __init__(self, api_key: str | None = None):
        if Cerebras is None:
            raise ImportError("cerebras-cloud-sdk not installed")
        self.client = Cerebras(api_key=api_key or os.environ.get("CEREBRAS_API_KEY"))
        CEREBRAS_API_KEY="csk-99k6tptdvethm46ey345fcvvpn3k55thcdjk9c9dj2r4e622"

    def parse(self, instruction: str) -> List[dict]:
        """Convert instruction to a list of structured operations."""
        msg = [{"role": "user", "content": instruction}]
        resp = self.client.chat.completions.create(
            model=self.MODEL_NAME,
            messages=msg,
        )
        content = resp.choices[0].message.content  # type: ignore[attr-defined]
        try:
            return json.loads(content)
        except Exception:  # pragma: no cover
            # Fallback: wrap raw text
            return [{"op": "raw", "content": content}]


class LogicEngine:
    """Execute structured operations on a matrix."""

    def __init__(self, matrix: Matrix, address_map: AddressMap):
        self.matrix = matrix
        self.addr_map = address_map

    def apply(self, ops: Iterable[dict]) -> None:
        for op in ops:
            self._apply_single(op)

    # ------------------------------------------------------------------
    def _apply_single(self, op: dict) -> None:  # noqa: C901 (simple placeholder)
        name = op.get("op")
        if name == "raw":
            # placeholder – nothing to do
            return
        if name == "add":
            self._binary(op, lambda a, b: a + b)
        elif name == "subtract":
            self._binary(op, lambda a, b: a - b)
        else:
            raise ValueError(f"Unsupported op: {name}")

    def _binary(self, op: dict, fn):
        tgt = op["target"]
        left = self._resolve(op["left"])
        right = self._resolve(op["right"])
        res = fn(left, right)
        r, c = self._address_to_idx(tgt)
        # ensure row exists
        while r >= len(self.matrix):
            self.matrix.append([])
        row = self.matrix[r]
        while c >= len(row):
            row.append(None)
        row[c] = res

    # ------------------------------------------------------------------
    def _resolve(self, ref: Union[str, float, int]) -> Any:
        if isinstance(ref, (int, float)):
            return ref
        r, c = self._address_to_idx(ref)
        return self.matrix[r][c]

    def _address_to_idx(self, addr: str) -> Tuple[int, int]:
        # very simple conversion (supports AA, AB …)
        col_part = "".join(filter(str.isalpha, addr)).upper()
        row_part = "".join(filter(str.isdigit, addr))
        col_num = 0
        for ch in col_part:
            col_num = col_num * 26 + (ord(ch) - 64)
        return int(row_part) - 1, col_num - 1


class Exporter:
    """Write matrix back to CSV."""

    def __init__(self, matrix: Matrix):
        self.matrix = matrix

    def to_csv(self, path: str) -> None:
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerows(self.matrix)


__all__ = [
    "SpreadsheetLoader",
    "AiParser",
    "LogicEngine",
    "Exporter",
    "read_workbook_as_matrices",
    "ask_question_about_workbook",
    "ai_orchestrate_and_write_numbers",
    "ai_question_compute_from_workbook",
    "ai_plan_formulas_from_workbook",
    "evaluate_formulas_to_values",
    "ai_formulas_to_numbers_and_write",
    "attach_cell_addresses",
    "list_assumptions",
]

def read_workbook_as_matrices(path: str) -> Dict[str, Matrix]:
    """Load all sheets from an Excel workbook into a dict of 2-D matrices."""
    wb = openpyxl.load_workbook(path, data_only=True)
    data: Dict[str, Matrix] = {}
    for ws in wb.worksheets:
        sheet_matrix: Matrix = []
        for row in ws.iter_rows(values_only=True):
            sheet_matrix.append(list(row))
        data[ws.title] = sheet_matrix
    return data


def ask_question_about_workbook(
    question: str,
    sheets: Dict[str, Matrix],
    api_key: str | None = None,
) -> str:
    """Ask the Cerebras model a question about all sheets' data.

    The entire workbook (sheet -> matrix) is serialized to JSON and provided as
    context. Keep questions focused to avoid overly large prompts.
    """
    if Cerebras is None:
        raise ImportError("cerebras-cloud-sdk not installed")

    client = Cerebras(api_key=api_key or os.environ.get("CEREBRAS_API_KEY"))
    system_ctx = (
        "You are a data analyst. Answer questions about this spreadsheet. "
        "Be concise and perform calculations as needed. Data:\n" + json.dumps(sheets)
    )
    messages = [
        {"role": "system", "content": system_ctx},
        {"role": "user", "content": question},
    ]
    resp = client.chat.completions.create(
        model=AiParser.MODEL_NAME,
        messages=messages,
    )
    return resp.choices[0].message.content  # type: ignore[attr-defined]


# ---------------------------------------------------------------------------
# AI-orchestrated numbers (no formulas):
# 1) Extract Assumptions + 2022 base from Operating Model
# 2) Ask AI for numeric updates per row_label and year
# 3) Validate quick sanity, then write values to Operating Model E:J
# ---------------------------------------------------------------------------

def _find_year_headers(ws) -> Dict[int, int]:
    """Return mapping of year -> column index by scanning header area.
    Accepts numeric or string years; attempts a fallback contiguous mapping.
    """
    def parse_year(v) -> int | None:
        if isinstance(v, (int, float)):
            y = int(v)
            return y if 2000 <= y <= 2100 else None
        if isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                y = int(s)
                return y if 2000 <= y <= 2100 else None
        return None

    max_scan_rows = min(25, ws.max_row)
    # First pass: collect any year-like headers by row
    for header_row in range(1, max_scan_rows + 1):
        row_vals = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
        candidates: Dict[int, int] = {}
        for c_idx, val in enumerate(row_vals, start=1):
            y = parse_year(val)
            if y is not None:
                candidates[y] = c_idx
        if len(candidates) >= 3:
            return dict(sorted(candidates.items()))

    # Fallback: find a single 2022 header and assume contiguous to the right
    for header_row in range(1, max_scan_rows + 1):
        for c in range(1, ws.max_column + 1):
            y = parse_year(ws.cell(row=header_row, column=c).value)
            if y == 2022:
                # Map out up to 6 years (2022..2027)
                mapping: Dict[int, int] = {}
                year = 2022
                col = c
                while col <= ws.max_column and year <= 2100:
                    mapping[year] = col
                    year += 1
                    col += 1
                if len(mapping) >= 3:
                    return mapping
    return {}


def _find_row_by_label(ws, label: str, label_col: int = 3) -> int:
    """Find row index in ws where column label_col equals label."""
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=label_col).value).strip() == label:
            return r
    raise KeyError(f"Row label not found: {label}")


def _extract_assumptions(path: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Assumptions"]
    year_cols = _find_year_headers(ws)
    def row_vals(lbl: str) -> Dict[int, float]:
        r = _find_row_by_label(ws, lbl)
        out: Dict[int, float] = {}
        for year, col in year_cols.items():
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            out[year] = float(v)
        return out
    return {
        "revenue_growth": row_vals("Revenue (Yoy.)"),
        "cogs_increasing_margin": row_vals("COGS (increasing marging)"),
        "sga_pct": row_vals("SG&A (% of Rev.)"),
        "da_pct": row_vals("D&A (% of Rev.)"),
        "tax_rate": row_vals("Tax"),
        "inventory_days": row_vals("Inventory Days"),
        "debtor_days": row_vals("Debtor Days"),
        "creditor_days": row_vals("Creditor Days"),
        "capex_pct": row_vals("CapEx (% of Rev.)"),
        "lt_rate": row_vals("Long-Term Debt Interest"),
        "revolver_rate": row_vals("Revolver Interest"),
        "mandatory_amort": row_vals("Debt redemption (mandatory)"),
    }



def _extract_operating_base_2022(path: str) -> Dict[str, float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Operating Model"]
    year_cols = _find_year_headers(ws)
    if 2022 not in year_cols:
        raise ValueError("2022 column not found in Operating Model header")
    col_2022 = year_cols[2022]
    def v(lbl: str) -> float:
        r = _find_row_by_label(ws, lbl)
        val = ws.cell(row=r, column=col_2022).value
        return 0.0 if val is None else float(val)
    return {
        "Revenue": v("Revenue"),
        "COGS": v("COGS"),
        "SG&A": v("SG&A"),
        "D&A": v("D&A"),
        "Interest Income": v("Interest Income"),
        "Interest Expense": v("Interest Expense"),
        "PP&E": v("PP&E"),
        "Inventory": v("Inventory"),
        "Accounts Receivable": v("Accounts Receivable"),
        "Accounts Payable": v("Accounts Payable"),
        "Cash": v("Cash"),
        "Long-Term Debt": v("Long-Term Debt"),
        "Shareholder's Equity": v("Shareholder's Equity"),
    }


def _ai_plan_numeric_updates(path: str, question: str | None = None) -> Dict[str, Any]:
    """Ask AI for numeric updates to Operating Model. Returns plan dict:
    {"updates": [{"row_label": str, "year_values": {"2023": num, ...}}]}
    """
    if Cerebras is None:
        raise ImportError("cerebras-cloud-sdk not installed")
    assumptions = _extract_assumptions(path)
    base2022 = _extract_operating_base_2022(path)
    client = Cerebras(api_key=os.environ.get("CEREBRAS_API_KEY"))
    system = (
        "You are a financial modeling assistant. Based on the given assumptions and 2022 base, "
        "produce NUMERIC values (no formulas) for the Operating Model rows for years 2023-2027. "
        "Output strict JSON: {\"updates\":[{\"row_label\":str,\"year_values\":{\"2023\":num,...,\"2027\":num}}]}. "
        "Only include rows that should be computed (e.g., Gross Profit, margins, EBIT, PBT, Taxes, Net Income, Δ NWC, Capex, Pre-Financing Cash Flow, LT Debt BoP/Repay/EoP as implied, Revolver BoP/Draw/EoP if needed, Cash EoB)."
    )
    user = {
        "assumptions": assumptions,
        "base_2022": base2022,
        "note": question or "Complete the Operating Model values for 2023-2027."
    }
    resp = client.chat.completions.create(
        model=AiParser.MODEL_NAME,
        messages=[{"role": "system", "content": json.dumps({"instruction": system})},
                  {"role": "user", "content": json.dumps(user)}],
    )
    content = resp.choices[0].message.content  # type: ignore[attr-defined]
    try:
        plan = json.loads(content)
        if not isinstance(plan, dict) or "updates" not in plan:
            raise ValueError("Invalid plan schema")
        return plan
    except Exception as e:
        raise ValueError(f"AI plan parse error: {e}")


def _write_operating_updates(path: str, plan: Dict[str, Any]) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["Operating Model"]
    year_cols = _find_year_headers(ws)
    for upd in plan.get("updates", []):
        label = upd.get("row_label")
        yvals: Dict[str, Any] = upd.get("year_values", {})
        if not label or not isinstance(yvals, dict):
            continue
        try:
            r = _find_row_by_label(ws, label)
        except KeyError:
            continue
        for y_str, val in yvals.items():
            try:
                y = int(y_str)
            except Exception:
                continue
            if y not in year_cols:
                continue
            if val is None:
                ws.cell(row=r, column=year_cols[y]).value = None
            else:
                ws.cell(row=r, column=year_cols[y]).value = float(val)
    wb.save(path)


def ai_orchestrate_and_write_numbers(excel_path: str, question: str | None = None) -> Dict[str, Any]:
    """High-level entry: get AI numeric plan and write values into Operating Model."""
    plan = _ai_plan_numeric_updates(excel_path, question)
    _write_operating_updates(excel_path, plan)
    return plan


def ai_question_compute_from_workbook(excel_path: str, question: str, write: bool = False) -> Dict[str, Any]:
    """AI-inspects the entire workbook to compute numbers for Operating Model.

    Input: full workbook matrices + question. Output: plan with updates and optional answer.
    If write=True, values are written into Operating Model E:J.
    """
    if Cerebras is None:
        raise ImportError("cerebras-cloud-sdk not installed")
    sheets = read_workbook_as_matrices(excel_path)
    client = Cerebras(api_key=os.environ.get("CEREBRAS_API_KEY"))
    system = (
        "You are a financial modeling assistant. Using the provided workbook data, "
        "infer where assumptions and base values live, compute the required Operating Model line items "
        "for years 2023-2027, and return STRICT JSON only."
        " Schema: {\"answer\": str, \"updates\": [{\"row_label\": str, \"year_values\": {\"2023\": num, ..., \"2027\": num}}]}"
        " Use the sheet named 'Operating Model' for placements; its column C contains row labels, and columns E:J correspond to years 2022-2027."
        " Provide NUMBERS only (no formulas) under year_values."
    )
    payload = {"workbook": sheets, "question": question}
    resp = client.chat.completions.create(
        model=AiParser.MODEL_NAME,
        messages=[
            {"role": "system", "content": json.dumps({"instruction": system})},
            {"role": "user", "content": json.dumps(payload)},
        ],
    )
    content = resp.choices[0].message.content  # type: ignore[attr-defined]
    plan = json.loads(content)
    if write:
        _write_operating_updates(excel_path, plan)
    return plan


# ---------------------------------------------------------------------------
# AI-planned formulas (DSL) → Python-evaluated numbers → optional write
# ---------------------------------------------------------------------------

def ai_plan_formulas_from_workbook(excel_path: str, question: str) -> Dict[str, Any]:
    """Ask AI to produce a formula plan (JSON DSL) keyed by row_label.

    DSL terminals:
      - {"const": number}
      - {"ref": "Revenue"}  # reference a line item same-year
      - {"series": "revenue_growth"}  # assumption series by year
      - {"op":"lag", "ref":"Revenue", "periods": 1}

    DSL ops (n-ary unless noted): add, sub, mul, div, neg (unary).

    Returns: {"formulas": [{"row_label": str, "expr": <dsl_expr>}, ...]}
    """
    if Cerebras is None:
        raise ImportError("cerebras-cloud-sdk not installed")
    sheets = read_workbook_as_matrices(excel_path)
    assumptions = _extract_assumptions(excel_path)
    base2022 = _extract_operating_base_2022(excel_path)
    client = Cerebras(api_key=os.environ.get("CEREBRAS_API_KEY"))
    system = (
        "You are a financial modeling assistant. Generate a JSON DSL of formulas that, when evaluated, "
        "will populate 2023-2027 values for key Operating Model rows. Use only the allowed DSL ops. "
        'Return STRICT JSON only with schema: {"formulas": [{"row_label": "string", "expr": "dsl"}]}'
    )
    payload = {
        "workbook": sheets,
        "assumptions": assumptions,
        "base_2022": base2022,
        "question": question,
        "dsl_spec": {
            "terminals": ["const", "ref", "series", "lag"],
            "ops": ["add", "sub", "mul", "div", "neg"],
            "notes": "Costs should be negative (e.g., COGS, SG&A, D&A). Use lag(ref,1) for prior-year linkage."
        },
    }
    resp = client.chat.completions.create(
        model=AiParser.MODEL_NAME,
        messages=[
            {"role": "system", "content": json.dumps({"instruction": system})},
            {"role": "user", "content": json.dumps(payload)},
        ],
    )
    content = resp.choices[0].message.content  # type: ignore[attr-defined]
    plan = json.loads(content)
    if not isinstance(plan, dict) or "formulas" not in plan:
        raise ValueError("Invalid formula plan schema")
    return plan


def evaluate_formulas_to_values(
    excel_path: str,
    formula_plan: Dict[str, Any],
    years: Tuple[int, ...] = (2023, 2024, 2025, 2026, 2027),
) -> Dict[str, Any]:
    """Evaluate the AI DSL deterministically to numeric updates.

    Returns {"updates": [{"row_label": str, "year_values": {str(year): number}}]}
    """
    assumptions = _extract_assumptions(excel_path)
    base2022 = _extract_operating_base_2022(excel_path)

    values: Dict[str, Dict[int, float]] = {}
    # Seed 2022 for refs
    for k, v in base2022.items():
        values.setdefault(k, {})[2022] = v

    # Helper to get series val
    def series_val(name: str, year: int):
        s = assumptions.get(name, {})
        v = s.get(year)
        return None if v is None else float(v)

    # DSL evaluator (single expression at a given year)
    def eval_expr(expr: Any, year: int):
        if isinstance(expr, dict):
            if "const" in expr:
                return float(expr["const"])
            if "ref" in expr:
                lbl = str(expr["ref"])  # same-year reference
                val = values.get(lbl, {}).get(year)
                return None if val is None else float(val)
            if "series" in expr:
                return series_val(str(expr["series"]), year)
            op = expr.get("op")
            if op == "lag":
                ref = str(expr.get("ref"))
                periods = int(expr.get("periods", 1))
                val = values.get(ref, {}).get(year - periods)
                return None if val is None else float(val)
            if op in ("add", "sub", "mul", "div"):
                args = expr.get("args", [])
                nums = [eval_expr(a, year) for a in args]
                # If any term is None, result remains None (placeholder)
                if any(n is None for n in nums):
                    return None
                if op == "add":
                    return float(sum(nums))
                if op == "sub":
                    if not nums:
                        return None
                    acc = nums[0]
                    for n in nums[1:]:
                        acc -= n
                    return float(acc)
                if op == "mul":
                    acc = 1.0
                    for n in nums:
                        acc *= n
                    return float(acc)
                if op == "div":
                    if not nums:
                        return None
                    acc = nums[0]
                    for n in nums[1:]:
                        if n in (None, 0):
                            return None
                        acc = acc / n
                    return float(acc)
            if op == "neg":
                inner = eval_expr(expr.get("arg"), year)
                return None if inner is None else -float(inner)
        # Fallback
        return None

    # Normalize plan: parse if string; accept direct 'updates' passthrough
    if isinstance(formula_plan, str):
        try:
            formula_plan = json.loads(formula_plan)
        except Exception:
            return {"updates": []}
    if isinstance(formula_plan, dict) and "updates" in formula_plan and "formulas" not in formula_plan:
        # Already computed updates
        return {"updates": formula_plan.get("updates", [])}

    formulas = formula_plan.get("formulas", []) if isinstance(formula_plan, dict) else []
    if isinstance(formulas, str):
        try:
            formulas = json.loads(formulas)
        except Exception:
            formulas = []
    # Iterate passes to resolve dependencies
    for _ in range(5):  # limited passes
        changed = False
        for f in formulas:
            if not isinstance(f, dict):
                continue
            row_label = str(f.get("row_label")) if f.get("row_label") is not None else ""
            expr = f.get("expr")
            if not row_label or expr is None:
                continue
            for y in years:
                new_val = eval_expr(expr, y)
                cur_val = values.get(row_label, {}).get(y)
                # Update if newly available or changed numerically
                if cur_val is None and new_val is not None:
                    values.setdefault(row_label, {})[y] = new_val
                    changed = True
                elif (cur_val is not None and new_val is not None
                      and abs(new_val - cur_val) > 1e-9):
                    values.setdefault(row_label, {})[y] = new_val
                    changed = True
        if not changed:
            break

    updates = []
    for row_label, by_year in values.items():
        year_values = {str(y): by_year.get(y) for y in years if y in by_year}
        if year_values:
            updates.append({"row_label": row_label, "year_values": year_values})
    return {"updates": updates}


def ai_formulas_to_numbers_and_write(
    excel_path: str,
    question: str,
    write: bool = True,
) -> Dict[str, Any]:
    """End-to-end: AI formulas → Python numbers → optional write → return updates."""
    plan = ai_plan_formulas_from_workbook(excel_path, question)
    updates = evaluate_formulas_to_values(excel_path, plan)
    if write:
        _write_operating_updates(excel_path, updates)
    return updates


def attach_cell_addresses(excel_path: str, updates: Dict[str, Any]) -> Dict[str, Any]:
    """Return updates enriched with the exact Excel cell addresses for each year.

    Output schema mirrors input updates and adds:
      {"updates": [{..., "cell_addresses": {"2023": "F9", ...}}]}
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Operating Model"]
    year_cols = _find_year_headers(ws)
    enriched: List[Dict[str, Any]] = []
    for upd in updates.get("updates", []):
        if not isinstance(upd, dict):
            continue
        row_label = upd.get("row_label")
        yvals = upd.get("year_values", {}) if isinstance(upd.get("year_values"), dict) else {}
        if not isinstance(row_label, str):
            continue
        try:
            r = _find_row_by_label(ws, row_label)
        except KeyError:
            continue
        cell_addrs: Dict[str, str] = {}
        for y_str in yvals.keys():
            try:
                y = int(y_str)
            except Exception:
                continue
            col = year_cols.get(y)
            if col is None:
                continue
            cell_addrs[y_str] = ws.cell(row=r, column=col).coordinate
        entry = dict(upd)
        entry["cell_addresses"] = cell_addrs
        enriched.append(entry)
    return {"updates": enriched}


def list_assumptions(excel_path: str) -> List[Dict[str, Any]]:
    """Flatten assumptions to a simple list of {name, year, value} for printing/logging."""
    a = _extract_assumptions(excel_path)
    flat: List[Dict[str, Any]] = []
    for name, by_year in sorted(a.items()):
        for year in sorted(by_year.keys()):
            flat.append({"name": name, "year": int(year), "value": by_year[year]})
    return flat

