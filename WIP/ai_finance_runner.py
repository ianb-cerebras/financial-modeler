"""ai_finance_runner.py

Streamlined entry-point for asking Cerebras financial-model questions.
Workflow:
1. Load specified sheets from an Excel workbook and serialise to JSON.
2. Prompt the Cerebras model with the workbook JSON + user question.
   The model must respond with strict JSON:
   {
     "answer_label": str,        # which Operating Model row holds the answer
     "formulas": [               # DSL formulas to compute rows 2023-2027
       {"row_label": str, "expr": <dsl_expr>},
       ...
     ]
   }
   DSL terminals:
     {"const": num}
     {"ref": "Revenue"}
     {"series": "revenue_growth"}
     {"op":"lag","ref":"Revenue","periods":1}
   DSL ops: add, sub, mul, div, neg
3. Deterministically evaluate the formulas for years 2023-2027.
4. Optionally write numeric results back into the Operating Model sheet.
5. Print the requested answer row or the entire update block.

Note: This is a *skeleton* – fill in TODOs for production use.
"""
from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from typing import Dict, Any, Tuple

import openpyxl

# -- You already have these utilities in spreadsheet_pipeline -----------------
from runner_core import (
    load_all_sheets,
    list_assumptions,
    call_cerebras_for_plan,
    evaluate_dsl,
    validate_balance_sheet,
)

from spreadsheet_pipeline import (
    attach_cell_addresses,  # reuse writer utilities for now
    _write_operating_updates,
)

try:
    from cerebras.cloud.sdk import Cerebras  # type: ignore
except ImportError:
    Cerebras = None  # type: ignore

YEARS: Tuple[int, ...] = (2023, 2024, 2025, 2026, 2027)

# ---------------------------------------------------------------------------
# Cerebras call
# ---------------------------------------------------------------------------

def call_cerebras_for_plan(workbook_json: Dict[str, Any], question: str) -> Dict[str, Any]:
    """Send workbook and question to Cerebras, return JSON plan."""
    if Cerebras is None:
        raise ImportError("cerebras-cloud-sdk not installed")
    client = Cerebras(api_key=os.environ.get("CEREBRAS_API_KEY"))
    CEREBRAS_API_KEY="csk-99k6tptdvethm46ey345fcvvpn3k55thcdjk9c9dj2r4e622"

    system_prompt = (
        "You are a financial‐model assistant. Using the provided workbook JSON, "
        "decide which formulas are required to answer the user's question. "
        "Output STRICT JSON only with schema: {answer_label, formulas}. "
        "DSL terminals: const, ref, series, lag. Ops: add, sub, mul, div, neg. "
        "Use column‐C labels from 'Operating Model' for row_label."
    )
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps({"workbook": workbook_json, "question": question})},
    ]
    resp = client.chat.completions.create(model="qwen-3-coder-480b", messages=messages)
    content = resp.choices[0].message.content  # type: ignore[attr-defined]
    return json.loads(content)


# ---------------------------------------------------------------------------
# CLI helpers
# ---------------------------------------------------------------------------

def load_selected_sheets(path: str, sheets: Tuple[str, ...] | None = None) -> Dict[str, Any]:
    """Return JSON-ready dict of sheet -> 2-D matrix.

    If *sheets* is None, load every sheet in the workbook; otherwise load
    only the specified names, silently skipping missing ones.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    targets = wb.sheetnames if sheets is None else list(sheets)
    out: Dict[str, Any] = {}
    for name in targets:
        if name in wb.sheetnames:
            ws = wb[name]
            out[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Ask a financial question about an Excel model via Cerebras.")
    parser.add_argument("excel", help="Path to workbook (.xlsx)")
    parser.add_argument("question", help="Financial question in natural language")
    parser.add_argument("--write", action="store_true", help="Write computed numbers back into Operating Model")
    parser.add_argument("--print-assumptions", action="store_true", help="Print flattened assumptions and exit")
    parser.add_argument("--api-key", dest="api_key", default=None, help="Cerebras API key (overrides CEREBRAS_API_KEY env var)")
    parser.add_argument("--debug", action="store_true", help="Print assumptions, AI plan, and computed updates")
    parser.add_argument("--log", action="store_true", help="Save prompt, plan, updates to logs/run_*.json")
    args = parser.parse_args()

    # Optional: print assumptions and exit
    if args.print_assumptions:
        print(json.dumps(list_assumptions(args.excel), indent=2))
        return

    # 1. Load all sheets for context
    wb_json = load_all_sheets(args.excel)

    # 2. Call Cerebras for formula plan (DSL mode)
    plan = call_cerebras_for_plan(wb_json, args.question)
    if args.debug:
        print("=== ASSUMPTIONS ===")
        print(json.dumps(list_assumptions(args.excel), indent=2))
        print("\n=== AI PLAN (formulas) ===")
        # Clean formula strings to remove any computed answers like "= 123.45"
        plan_to_print: Dict[str, Any] = dict(plan) if isinstance(plan, dict) else {"raw": plan}
        formulas = plan_to_print.get("formulas") if isinstance(plan_to_print, dict) else None
        if isinstance(formulas, list):
            cleaned: list[Any] = []
            for item in formulas:
                # If the model returned human-readable strings like
                # "Revenue: 950 * (1 + 0.05) = 997.5", strip the trailing
                # computed part so only the formula remains.
                if isinstance(item, str):
                    s = item.strip()
                    # Split on '=' and keep the left side
                    if "=" in s:
                        s = s.split("=", 1)[0].strip()
                    # If there's a leading label like "Label: expr", strip the label
                    if ":" in s:
                        left, right = s.split(":", 1)
                        # Heuristic: if left looks like a label (no operators), drop it
                        if not any(ch in left for ch in "+-*/()"):
                            s = right.strip()
                    cleaned.append(s)
                else:
                    cleaned.append(item)
            plan_to_print["formulas"] = cleaned
        print(json.dumps(plan_to_print, indent=2))

    # 3. Evaluate DSL deterministically
    assumptions_flat = list_assumptions(args.excel)
    updates = evaluate_dsl(plan, wb_json, assumptions_flat, YEARS)
    if args.debug:
        print("\n=== COMPUTED UPDATES (numbers) ===")
        print(json.dumps(updates, indent=2))
    enriched = attach_cell_addresses(args.excel, updates)
    if args.debug:
        print("\n=== ENRICHED WITH CELL ADDRESSES ===")
        print(json.dumps(enriched, indent=2))

    # 4. Validate balance sheet before write
    ok, report = validate_balance_sheet(enriched)
    if not ok:
        print("\nValidation FAILED: Balance Sheet does not balance")
        print(json.dumps(report, indent=2))
        if args.write:
            print("Write aborted due to validation failure.")
    else:
        if args.debug:
            print("\nBalance Sheet validation OK.")
        if args.write:
            _write_operating_updates(args.excel, enriched)
            print("Values written to workbook.")

    # 5. Print answer row if provided
    answer_label = plan.get("answer_label") or plan.get("answer")
    if answer_label:
        ans = next((u for u in enriched["updates"] if u.get("row_label") == answer_label), None)
        if ans is not None:
            print("Answer:", json.dumps(ans, indent=2))
        else:
            print(json.dumps(enriched, indent=2))
    else:
        print(json.dumps(enriched, indent=2))

    # 6. Optional logging
    if args.log:
        logs_dir = Path("logs")
        logs_dir.mkdir(exist_ok=True)
        import datetime, uuid
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = logs_dir / f"run_{ts}_{uuid.uuid4().hex[:6]}.json"
        payload = {
            "question": args.question,
            "assumptions": list_assumptions(args.excel),
            "plan": plan,
            "updates": enriched,
            "validation": report,
        }
        fname.write_text(json.dumps(payload, indent=2))
        if args.debug or args.log:
            print(f"Run logged to {fname}")


if __name__ == "__main__":
    main()

