import os
import json
from typing import Dict, Any, List, Optional
import openpyxl
from cerebras.cloud.sdk import Cerebras
import time
import logging
import random
import re

# ---------------- User configurable paths ----------------
SOURCE_FILE = "/Users/ianbaime/Desktop/dummydata.xlsx"  # primary data workbook to read and write

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Global Cerebras client & model configuration
# ---------------------------------------------------------------------------
CLIENT = Cerebras(api_key=os.environ.get("CEREBRAS_API_KEY"))
MODEL = "qwen-3-coder-480b"

# ---------------------------------------------------------------------------
# Spreadsheet helpers
# ---------------------------------------------------------------------------
def load_excel_data(filepath: str) -> Dict[str, List[List[Any]]]:
    """Load every worksheet in an Excel workbook into a matrix list-of-lists."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data: Dict[str, List[List[Any]]] = {}
    for ws in wb.worksheets:
        sheet_matrix = [list(row) for row in ws.iter_rows(values_only=True)]
        data[ws.title] = sheet_matrix
        logger.debug("Loaded sheet '%s' with %d rows", ws.title, len(sheet_matrix))
    return data

# ---------------------------------------------------------------------------
# Cerebras wrapper with retry/backoff
# ---------------------------------------------------------------------------
def _chat_with_retries(messages: List[Dict[str, Any]], *, purpose: str, max_retries: int = 3, base_delay: float = 0.5) -> str:
    """Call Cerebras chat API with exponential-backoff retries."""
    last_error: Exception | None = None
    for attempt in range(1, max_retries + 1):
        try:
            logger.info("Cerebras call (%s) attempt %d/%d", purpose, attempt, max_retries)
            response = CLIENT.chat.completions.create(model=MODEL, messages=messages)
            return response.choices[0].message.content
        except Exception as e:
            last_error = e
            if attempt == max_retries:
                break
            sleep_s = base_delay * (2 ** (attempt - 1)) + random.uniform(0, 0.25)
            logger.warning("Cerebras call failed (%s) on attempt %d: %s – retrying in %.2fs", purpose, attempt, e, sleep_s)
            time.sleep(sleep_s)
    logger.error("Cerebras call failed after %d attempts (%s)", max_retries, purpose)
    raise last_error if last_error else RuntimeError("Unknown Cerebras error")

# ---------------------------------------------------------------------------
# AI-assisted analytics steps
# ---------------------------------------------------------------------------
def identify_assumptions(workbook_data: Dict[str, Any], core_elements: Optional[str] = None, question: Optional[str] = None) -> str:
    """Cerebras call – extract numeric data & assumptions from raw workbook."""
    # ------------------------------------------------------------------
    # The model’s job is ONLY to find the raw inputs (assumptions) that
    # the hard-coded forecast engine will later consume.  We therefore ask
    # for a single JSON object named "assumptions" whose keys exactly
    # match the list agreed with the user.
    # ------------------------------------------------------------------
    system_prompt = (
        "You are a financial analyst. Your ONLY task: scan the raw workbook data and "
        "return a JSON object named assumptions that contains the following keys – ALL in lower-case, "
        "spaces replaced with underscores, and units in plain numbers (no formatting):\n\n"
        "  • base_revenue                    # starting revenue dollar amount\n"
        "  • revenue_growth_rate             # as decimal e.g. 0.05 for 5 %\n"
        "  • base_cogs_percent              # starting COGS % of revenue (decimal)\n"
        "  • cogs_delta_per_year            # annual increase in COGS % (decimal)\n"
        "  • sg&a_percent                     # decimal (negative not required)\n"
        "  • d&a_percent                      # decimal (negative not required)\n"
        "  • average_cash_balance\n"
        "  • average_debt_balance\n"
         " • EBIT\n"
        "  • cash_interest_rate               # decimal\n"
        "  • debt_interest_rate               # decimal\n"
        "  • effective_tax_rate               # decimal\n\n"
        "Guidance:\n"
        "• The source workbook may label these drivers differently. Map synonymous labels to the keys above.\n"
        "    – Map any of: 'cash', 'cash balance', 'cash & equivalents', 'average cash', 'cash bal' to average_cash_balance.\n"
        "    – Map any of: 'debt', 'total debt', 'average debt', 'lt debt', 'long-term debt', 'revolver balance' to average_debt_balance.\n"
        "    – If interest rates appear by instrument (e.g., 'revolver rate', 'term loan rate'), choose the main corporate borrowing rate for debt_interest_rate.\n"
        "• If the workbook already shows a calculated Interest Expense line, back-solve the implied average_debt_balance = interest_expense / debt_interest_rate.\n"
        "• Use the best available numeric evidence; if multiple candidates exist, prefer the one closest to the current year.\n\n"
        "Rules:\n"
        "• If a value is missing, set it to 0 (do NOT fabricate).\n"
        "• Return ONLY valid JSON. No commentary or code fences.\n"
    )
    context = {
        "workbook": workbook_data,
        "core_elements": core_elements,
        "question": question or "Extract the assumptions listed above from the workbook",
    }
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(context)},
    ]
    return _chat_with_retries(messages, purpose="identify_assumptions")


# ---------------------------------------------------------------------------
# Assumption parsing helper
# ---------------------------------------------------------------------------

def parse_assumptions(json_text: str) -> Dict[str, Any]:
    """Parse the assumptions JSON returned by the LLM.

    The LLM is instructed to output a single JSON object.  It may either be
    the assumptions object itself or an outer object with the key
    "assumptions".  This helper normalises both cases.
    """
    import json as _json

    try:
        data = _json.loads(json_text)
    except Exception as e:  # noqa: BLE001
        logger.error("Failed to parse assumptions JSON: %s", e)
        return {}

    # If the model wrapped the assumptions in an outer object, unwrap it.
    if isinstance(data, dict) and "assumptions" in data and isinstance(data["assumptions"], dict):
        return data["assumptions"]

    # Otherwise assume the dict itself *is* the assumptions mapping.
    if isinstance(data, dict):
        return data

    logger.error("Unexpected JSON structure for assumptions – expected dict, got %s", type(data).__name__)
    return {}


# ---------------------------------------------------------------------------
# Deterministic forecast engine
# ---------------------------------------------------------------------------

FORECAST_COLUMNS = ["B", "C", "D", "E", "F"]  # years 1-5 starting at column B


def _series_5y(base: float, growth: float) -> List[float]:
    """Return five-year series given base value and constant growth rate."""
    return [base * (1 + growth) ** i for i in range(5)]


def compute_forecast(
    assumptions: Dict[str, Any],
    *,
    rows_map: Optional[Dict[str, int]] = None,
    cols: Optional[List[str]] = None,
) -> Dict[str, Dict[str, float]]:
    """Compute 5-year P&L using hard-coded formulas guided by instructions.json.

    Args:
        assumptions: mapping returned by parse_assumptions(). Must contain the
            keys defined in the updated prompt.

    Returns:
        Nested dict: {sheet_name: {cell: value}}
    """
    # Extract inputs with defaults to 0 if missing.
    a = lambda k: float(assumptions.get(k, 0) or 0)  # noqa: E731

    base_revenue = a("base_revenue")
    growth_rate = a("revenue_growth_rate")

    base_cogs_pct = a("base_cogs_percent")
    cogs_delta = a("cogs_delta_per_year")
    sgna_pct = a("sg&a_percent")
    da_pct = a("d&a_percent")

    avg_cash = a("average_cash_balance")
    avg_debt = a("average_debt_balance")

    cash_rate = a("cash_interest_rate")
    debt_rate = a("debt_interest_rate")

    tax_rate = a("effective_tax_rate")

    # Revenue series – Year 1 is first projection (after assumptions)
    revenue_series = [base_revenue * (1 + growth_rate) ** (i + 1) for i in range(5)]

    # Cost of goods sold – COGS % increases by fixed delta each year starting in Year 1
    cogs_pct_series = [base_cogs_pct + cogs_delta * (i + 1) for i in range(5)]
    cogs_series = [rev * pct for rev, pct in zip(revenue_series, cogs_pct_series)]

    # Gross profit
    gp_series = [rev - cogs for rev, cogs in zip(revenue_series, cogs_series)]

    # SG&A and D&A (expenses as positive numbers for now)
    sgna_series = [rev * sgna_pct for rev in revenue_series]
    da_series = [rev * da_pct for rev in revenue_series]

    # EBIT
    ebit_series = [gp - sgna - da for gp, sgna, da in zip(gp_series, sgna_series, da_series)]

    # Interest income/expense
    int_income_series = [avg_cash * cash_rate for _ in range(5)]
    int_expense_series = [avg_debt * debt_rate for _ in range(5)]

    # Profit before tax (EBT)
    ebt_series = [ebit + ii - ie for ebit, ii, ie in zip(ebit_series, int_income_series, int_expense_series)]

    # Tax expense (0 if negative EBT)
    tax_series = [ebt * tax_rate if ebt > 0 else 0 for ebt in ebt_series]

    # Net income
    ni_series = [ebt - tax for ebt, tax in zip(ebt_series, tax_series)]

    # Build output dict
    sheet = "Income Statement"
    if rows_map is None:
        rows_map = LABEL_TO_ROW
    if cols is None:
        cols = FORECAST_COLUMNS
    out: Dict[str, Dict[str, float]] = {sheet: {}}

    series_map = {
        "revenue": revenue_series,
        "cogs": cogs_series,
        "gross profit": gp_series,
        "sg&a": sgna_series,
        "d&a": da_series,
        "ebit": ebit_series,
        "interest income": int_income_series,
        "interest expense": int_expense_series,
        "profit before tax": ebt_series,
        "tax expense": tax_series,
        "net income": ni_series,
    }

    for label, series in series_map.items():
        row_num = rows_map.get(label)
        if not row_num:
            continue
        for idx, col in enumerate(cols):
            cell = f"{col}{row_num}"
            out[sheet][cell] = series[idx]

    logger.info("Computed forecast with %d cells", len(out[sheet]))
    return out


# (Removed legacy JSON-to-values converter; not used in web flow)


# (Removed legacy template fill function; not used in web flow)


# ---------------------------------------------------------------------------
# Template sheet handling
# ---------------------------------------------------------------------------
# (Removed legacy template sheet copier; not used in web flow)


def populate_template(values: Dict[str, float], dest_path: str, sheet_name: str) -> None:
    """Write numbers into existing sheet, only where coordinates already exist."""
    wb = openpyxl.load_workbook(dest_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} not found in {dest_path}")
    ws = wb[sheet_name]

    for cell, num in values.items():
        try:
            ws[cell].value = num
        except ValueError:
            logger.debug("Skipping invalid cell coordinate %s", cell)

    wb.save(dest_path)
    logger.info("Populated %d values into sheet '%s'", len(values), sheet_name)


# ---------------------------------------------------------------------------
# Formula evaluation engine
# ---------------------------------------------------------------------------


# new function for inserting into any workbook
# (Removed legacy utility to create new sheet; web flow writes into existing sheet)


# ---------------------------------------------------------------------------
# Final AI review helper
# ---------------------------------------------------------------------------
def last_check(workbook_values: Dict[str, Dict[str, float]], summary_question: Optional[str] = None) -> str:
    """Ask the model to review the filled data and flag any issues.

    Args:
        workbook_values: Nested dict of the numbers we plan to write – typically the result of evaluate_formulas.
        summary_question: Optional extra question/prompts for the model.
    Returns:
        Model response containing either 'All good' or suggestions/corrections.
    """
    system_prompt = (
        "You are a formatting auditor. Examine the structured workbook data intended for presentation. "
        "Check for alignment with standard financial-model formatting: correct headers, year columns, sourceline totals, consistent decimal places, and logical ordering. "
        "If everything is presentation-ready, respond with 'PASS'. Otherwise, list the formatting issues and specify the exact cell edits required to fix them (sheet, cell, new_value or action)."
    )
    context = {
        "values": workbook_values,
        "question": summary_question or "Perform a final sanity check on the financial model output."
    }
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(context)},
    ]
    return _chat_with_retries(messages, purpose="final_review")

# ---------------------------------------------------------------------------
# Formatting auto-fix helper
# ---------------------------------------------------------------------------

def apply_formatting_fixes(review_text: str, workbook_path: str, sheet_name: str = "Income Statement") -> None:
    """Parse the LLM review for simple cell-level edits and apply them.

    Expected edit lines of the form:
      • SheetName | Cell | NewValue
    or
      • Cell | NewValue   (sheet inferred)

    Anything that cannot be parsed is ignored silently.
    If you see any data that is clearly wrong on the projection page, such as a missed decimal place, correct that too.
    This is deliberately conservative so the model can only tweak explicit cells.
    """
    import re
    edits = []
    pattern = re.compile(r"(?P<sheet>[A-Za-z0-9 _]+)?\s*\|\s*(?P<cell>[A-Za-z]+[0-9]+)\s*\|\s*(?P<value>.+)")
    for line in review_text.splitlines():
        m = pattern.search(line)
        if m:
            sheet = m.group("sheet") or sheet_name
            cell = m.group("cell").strip()
            new_val = m.group("value").strip()
            edits.append((sheet, cell, new_val))

    if not edits:
        print("No actionable formatting edits parsed from review.")
        return

    wb = openpyxl.load_workbook(workbook_path)
    applied = 0
    for sheet, cell, new_val in edits:
        if sheet not in wb.sheetnames:
            print(f"Skip unknown sheet {sheet} in edit {sheet}|{cell}")
            continue
        ws = wb[sheet]
        try:
            # attempt to convert numeric strings to float
            if re.fullmatch(r"-?\d+(\.\d+)?", new_val):
                ws[cell].value = float(new_val)
            else:
                ws[cell].value = new_val
            applied += 1
        except ValueError:
            print(f"Invalid cell coordinate in edit {sheet}|{cell}")

    if applied:
        wb.save(workbook_path)
        print(f"Applied {applied} formatting fixes as suggested by the model.")


# ---------------------------------------------------------------------------
# Sheet structure detection via LLM
# ---------------------------------------------------------------------------


TARGET_LABELS = [
    "revenue",
    "cogs",
    "gross profit",
    "sg&a",
    "d&a",
    "ebit",
    "interest income",
    "interest expense",
    "profit before tax",
    "tax expense",
    "net income",
]


def detect_structure(workbook: Dict[str, List[List[Any]]]) -> Dict[str, Any]: #designed specifically for 5 year projections
    """Ask the LLM to locate row numbers and first forecast column.

    Returns a dict {"start_column": "B", "rows": {label: row_num | None}}. 
    Raises RuntimeError if JSON cannot be parsed.
    """
    prompt = (
        "You are a spreadsheet auditor. Given a worksheet dump (list of rows) "
        "and a list of target P&L labels, find which ROW each label appears on.\n"
        "Return JSON exactly in this format:\n\n"
        "{\n  \"sheet_name\": \"<sheet>\",\n  \"start_column\": \"<letter>\",\n  \"rows\": { label: row_or_null }\n}\n\n"
        "Rules:\n"
        "• Match labels case-insensitively; ignore punctuation (%, &, ., -). Wording may differ.\n"
        "• Map common synonyms to target keys:\n"
        "    – 'Operating income' or 'Operating profit' → ebit\n"
        "    – 'COGS' or 'Cost of goods sold' → cogs\n"
        "    – 'SG&A' or 'Selling, general & administrative' → sg&a\n"
        "    – 'D&A' or 'Depreciation & amortization' → d&a\n"
        "    – 'EBT' or 'Profit before taxes' → profit before tax\n"
        "• If a label is missing, set its value to null.\n"
        "• start_column = the first completely empty column (rows 2-100 are empty)"
        "  for the matched rows (so forecasts won't overwrite historicals).\n"
        "• Do not include commentary – return only valid JSON."
    )

    context = {"workbook": workbook, "labels": TARGET_LABELS}

    messages = [
        {"role": "system", "content": prompt},
        {"role": "user", "content": json.dumps(context)},
    ]

    resp = _chat_with_retries(messages, purpose="detect_structure")

    try:
        data = json.loads(resp)
        assert all(k in data for k in ("sheet_name", "start_column", "rows"))
        return data
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"Failed to parse structure JSON: {e} – raw: {resp[:200]}") from e


def col_sequence(start_col: str, n: int = 5) -> List[str]:
    """Return a list of n sequential Excel column letters starting at start_col."""
    from openpyxl.utils import column_index_from_string, get_column_letter

    base_idx = column_index_from_string(start_col.upper())
    return [get_column_letter(base_idx + i) for i in range(n)]


# ---------------------------------------------------------------------------
# Command-line entrypoint
# ---------------------------------------------------------------------------
def main() -> None:
    """Run the forecasting pipeline writing results back into the source workbook."""
    user_question = input("Enter analysis question (or press Enter for default): ") or None

    import time as _time
    t_start = _time.perf_counter()

    wb_data = load_excel_data(SOURCE_FILE)

    # Ask the model to find which sheet holds the P&L along with the structure
    structure = detect_structure(wb_data)  # pass entire workbook
    sheet_name = structure.get("sheet_name", "Income Statement")

    # Load modeling rules so the LLM follows strict sign discipline
    core_rules: Optional[str] = None
    try:
        with open("instructions.json", "r") as f:
            core_rules = f.read()
    except Exception:
        core_rules = None

    print("\n--- Calling identify_assumptions ---")
    assumptions_text = identify_assumptions(wb_data, core_rules, user_question)
    print("AI assumptions raw response:\n", assumptions_text)

    assumptions = parse_assumptions(assumptions_text)
    print("\nParsed assumptions dict:\n", assumptions)

    print("\n--- Computing 5-year forecast locally ---")
    from openpyxl.utils import column_index_from_string, get_column_letter

    detected_col = structure["start_column"]
    start_idx = column_index_from_string(detected_col.upper())
    cols = [get_column_letter(start_idx + i) for i in range(5)]
    rows_map = {k: v for k, v in structure["rows"].items() if v}

    values = compute_forecast(assumptions, rows_map=rows_map, cols=cols)
    print("Forecast values dict:\n", values)

    review = last_check(values)
    print("\nAI review:\n", review)

    if "PASS" not in review.upper():
        apply_formatting_fixes(review, SOURCE_FILE, sheet_name)
        print("Re-ran formatting fixes based on model suggestions.")

    # Merge all values and populate
    combined: Dict[str, float] = {}
    for sheet_dict in values.values():
        combined.update(sheet_dict)

    populate_template(combined, SOURCE_FILE, sheet_name)
    t_end = _time.perf_counter()
    print("\nUpdated", SOURCE_FILE, "with formatted sheet", sheet_name)
    print(f"Total AI + projection runtime: {t_end - t_start:.2f} seconds")


if __name__ == "__main__":
    main()